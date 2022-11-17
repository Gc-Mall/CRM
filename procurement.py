"""
数据类别：品名、单价、数量、金额、供货商、日期
基础功能：数据导入，数据分析
进阶功能：同品类合并
数据安全: 数据校验 --> 数据数量校验
设计思路：数据结构分析，功能实现，调试与分析
"""

from datetime import date, datetime
import openpyxl as op
import re

import database

class Procurement():
    def __init__(self) -> None:
        pass


    def load_from_excel(self, excel_path) -> dict:
        """
        从excel模板中进行数据导入
        Data Structure: dic(dic)
        Parameter(name_list -> list): 菜品名字，用作data的key
        Parameter(info_detail -> dict): 菜品完整信息
        Parameter(data -> dict): 集大成者
        """
        name_list = []
        info_detail = {}
        data = {}
        workbook = op.load_workbook(excel_path)
        sheet = workbook["Sheet1"]
        #正则分析表格尺寸大小 -> 确定读取范围(B2:F?)
        size = re.search(r"(?<=:F)\d*", sheet.dimensions).group(0)
        #关于读取表格数据 -> Way 1: 指定A3; Way 2: 采用row, column
        #读取名称，存入name_list之中
        cell_name = sheet["B9:B"+size]
        for i in cell_name:
            for j in i:
                name_list.append(j.value)
        #读取详细信息，存取data之中
        for i in range(9, int(size)+1):
            #关于引用赋值的问题，需要注意，info_detail需置空
            info_detail = {}
            for j in range(2,7):
                if j == 2:
                    info_detail["name"] = sheet.cell(row = i, column = j).value
                if j == 3:
                    info_detail["quantity"] = sheet.cell(row = i, column = j).value
                elif j == 4:
                    info_detail["price"] = sheet.cell(row = i, column = j).value
                elif j == 5:
                    #计算及检测总金额功能
                    if sheet.cell(row = i, column = j).value == None:
                        info_detail["value"] = sheet.cell(row = i, column = 3).value * sheet.cell(row = i, column = 4).value
                else:
                    info_detail["supplier"] = sheet.cell(row = i, column = j).value
            data[name_list[i-9]] = info_detail
        #添加日期信息 --> 年月日
        data["date"] = datetime.date(sheet.cell(row = 2, column = 2).value)
        self.data_procurement = data
        return data


    def insert_into_mysql(self):
        """
        将采购信息导入至MySQL
        """
        database.insert_procurement(self.data_procurement)


    def export_from_mysql(self):
        """
        将采购信息从MySQL导出至excel模板进行分析
        """
        print()
    


i = Procurement()
i.load_from_excel(r"C:\Users\brian\Desktop\CRM\supplier.xlsx")
i.insert_into_mysql()
